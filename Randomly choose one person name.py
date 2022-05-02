#This app ranomly chooses one person to answer questions
#Out of cell A2 to A19 from excel file named hello that has all
#the student names from the class, we want ONE name to be randomly chosen.
from openpyxl import workbook
from openpyxl import load_workbook
import random
wb = load_workbook('hello.xlsx')
ws = wb.active
range = ws['A2':'A19']
l=[]
for cell in range:
     for x in cell:
         l.append(x.value)
print(l)
computer_action = random.choice(l)
print('The computer randomly chooses ' + computer_action + ' to answer the next question')